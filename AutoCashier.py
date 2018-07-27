from openpyxl import *	#for manipulating excel spreadsheet
import win32com.client as win32
import os
import Window as w

# get index of first row with entry
def findFirstEntryIndex(ws):
	#iterate through names of employees
	index = 0
	col = ws.Range("A1:A10")
	for cell in col:
		if not cell.Value:
			index +=1
			continue
		cellValue = cell.Value.split(',')
		if len(cellValue) == 2 and cellValue.pop(0) != "NAME\nLast Name":
			return index+1
		index += 1
		
	print ("Beginning not Found")
	return -1


# get index of last row with entry
def findLastEntryIndex(ws, index):
	i = 0 
	for cell in ws.Range("A1:A200"):
		if i < index:
			i+= 1
			continue
		else:
			i+=1
			if cell.Value == None or len(cell.Value.split(',')) != 2:
				break
	return i


def writeRow(ws, destRow, inputVal):
	for v in range(0,8):
		ws.cell(row=destRow,column=v+1).value=inputVal[v]


def main():
	#sets up work sheet. 
	fileName = 'AutoCashierTestWorkbook.xlsx'
	absPath = os.path.dirname(os.path.abspath(__file__))
	path = os.path.join(absPath, fileName)
	excel = win32.gencache.EnsureDispatch('Excel.Application')
	
	pwWindow = w.pwWindow()	#creates a password confirmation window

	wb = excel.Workbooks.Open(path, Password = pwWindow.getPassword())	#open workbook with password
	excel.Visible = True
	
	#generate sheet list
	sheetList = []
	for sh in wb.Sheets:
		sheetList.append(sh.Name)
	
	mWindow = w.mainWindow(sheetList)
	inputValues = mWindow.getUserInput()	#get formatted data
	ws = wb.Sheets(mWindow.location)		#set to proper location sheet
	
	indexFirst = findFirstEntryIndex(ws)
	print ("first: ", indexFirst)
	indexLast = findLastEntryIndex(ws, indexFirst)
	print ("last: ", indexLast)
	
	for i in inputValues:
		print (i[0])
	print (mWindow.location)

	"""
	# names that come first alphabetically are considered smaller
	print ("Length of loop: ", indexLast+len(inputValues))
	for inputVal in inputValues:
		for i in range(indexFirst,indexLast + len(inputValues)):	#accounts for entries that will be inputted
			currentCell = ws.cell(row=i,column=1).value	#gets name of current row
			nextCell = ws.cell(row=i+1, column=1).value #gets name of next row
			print (currentCell, "|", nextCell, "|", i)

			#insertion happens at the very beginning 
			if i == indexFirst and inputVal[0] < currentCell:
				print ("first insertion")
				ws.insert_rows(i,1) 						#insert blank row
				#ws.cell(row=i,column=1).value = inputVal[0]	#just write name column for now 
				writeRow(ws,i,inputVal)
				break
				
			#insertion happens at the end or somewhere in the middle
			if nextCell == None or (inputVal[0] > currentCell and inputVal[0] < nextCell):
				print ("\n	inserting between here!")
				ws.insert_rows(i+1,1) 							#insert blank row
				#ws.cell(row=i+1,column=1).value = inputVal[0]	#just write name column for now 
				writeRow(ws,i+1, inputVal)
				break
			
			print ()
	wb.save(fileName)
	"""


if __name__ == "__main__":
	main()
	